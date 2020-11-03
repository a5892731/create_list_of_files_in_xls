"""
create by: a5892731
date: 2020-11-01

"""

#import openpyxl
#import os
#from time import sleep
from os import listdir, getcwd, chdir
from openpyxl import Workbook
wb = Workbook()
ws = wb.active


bufor_file = []
bufor_folder = []
bufor_folders = []
bufor_column = []
#----------------------------------------------------------------------------------------------------------------------
#    files = os.listdir()
#    os.chdir('..')
#    direction = os.getcwd()
#-----------------------------------------------------------------------------------------------------------------------

def delete_unwanted_names(out=[]):
    if 'create_list_of_files_in_xls' in out:
        out.remove('create_list_of_files_in_xls')
    if '__pycache__' in out:
        out.remove('__pycache__')
    if '.idea' in out:
        out.remove('.idea')
    if '.git' in out:
        out.remove('.git')
    return out

def sort(direction=getcwd(), files=listdir(), column = 1):

    global bufor_file
    global bufor_folder
    global bufor_folders
    global bufor_column

    foolder_tree = []
    folders = []

    if files == []:
        return ''

    if 'list_creator.py' in files:
       chdir('..')
       direction = getcwd()
       files = listdir()

    files = delete_unwanted_names(files)


    for i in range(len(files)):
        if '.' in files[i]:
            bufor_file.append(files[i])
            bufor_folder.append(getcwd())
            #print('{} in {}'.format(bufor_file[-1], bufor_folder[-1]))
            bufor_column.append(column)

        else:
            chdir(files[i])
            column += 1

            sort(getcwd(), listdir(), column)

            bufor_folder.append(getcwd())
            bufor_column.append(column)

            bufor_folders.append(files[i])

            chdir('..')
            column -= 1

def print_folders(folders_list, line_xls):
    deeph = 0
    for folder in ((folders_list).strip(getcwd())).split('\\'):
        ws.cell(row=line_xls, column=2+ deeph, value=folder)
        deeph += 1

def printing(list1=[], list2=[], line_xls = 4, column_xls = 1):

    global bufor_column


    if list1 == []:
        return False

    print('-------------------------')

    ws.title = "Arkusz 1"

    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']

    for column in range(len(columns)):
        ws.column_dimensions[columns[column]].width = 40

    ws.append(('Lokalizacja folderu docelowego:', getcwd()))
    ws.append(('-----------------', '-----------------'))
    ws.append(('Plik:', ('Folder: ' + '> > >')))
    ws.cell(row=3, column=max(bufor_column)+1, value='Pełny adres:')

    for i in range(len(list1)):

            if list1[i] is not None:
                print('{} in {} >>>> Row:{}/Deeph:{}'.format(list1[i], list2[i], line_xls, bufor_column[i]-1))
                ws.cell(row=line_xls, column= 1, value=list1[i])
                print_folders(list2[i], line_xls)
                ws.cell(row=line_xls, column= max(bufor_column)+ 1, value=list2[i])
                line_xls += 1



# ----------------------------------------------------PROGRAM------------------------------------------------------


getcwd()

sort()

printing(bufor_file, bufor_folder)

chdir('create_list_of_files_in_xls')
wb.save('LIST OF FILES.xlsx')

wb.close()


print('-------------------------------------------------------------')
print('-------------------------------------------------------------')
input('Prees ENTER to EXIT')











